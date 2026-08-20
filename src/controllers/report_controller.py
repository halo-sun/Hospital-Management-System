"""Report controller – coordinates report generation and export."""
import logging
from datetime import date, datetime
from typing import Optional, Dict, Any, Tuple, List
from src.services.report_service import ReportService
from src.services.clinical_service import ClinicalService
from src.services.audit_service import AuditService
from src.constants import AuditAction, Role
from src.auth.rbac import require_role
from src.controllers.auth_controller import AuthController

logger = logging.getLogger(__name__)


class ReportController:
    """Handles report and analytics requests from the GUI layer."""

    def __init__(self, auth_ctrl: Optional[AuthController] = None) -> None:
        """Initialize ReportController with required services.

        Args:
            auth_ctrl: Auth controller providing the current role
                (used by the RBAC decorator).
        """
        self._auth_ctrl = auth_ctrl
        self._report_service = ReportService()
        self._clinical_service = ClinicalService()
        self._audit_service = AuditService()

    @property
    def _current_role(self) -> Optional[str]:
        """Return the logged-in user's role for RBAC checks."""
        if self._auth_ctrl is None:
            return None
        return self._auth_ctrl.current_role

    # ── Dashboard ──────────────────────────────────────────────

    @require_role(Role.ADMIN, Role.RECEPTIONIST)
    def get_dashboard_stats(self) -> Dict[str, Any]:
        """Get summary statistics for the admin dashboard.

        Returns:
            Dictionary with total_patients, total_doctors, etc.
        """
        return self._report_service.get_dashboard_stats()

    # ── Appointment Reports ────────────────────────────────────

    @require_role(Role.ADMIN)
    def get_daily_appointments(
        self, target_date: Optional[date] = None
    ) -> List[Dict[str, Any]]:
        """Get appointments for a specific day.

        Args:
            target_date: The date to report on.

        Returns:
            List of appointment records.
        """
        return self._report_service.get_daily_appointments(target_date)

    @require_role(Role.ADMIN)
    def get_monthly_appointments(
        self, year: Optional[int] = None, month: Optional[int] = None
    ) -> Dict[str, Any]:
        """Get monthly appointment statistics.

        Args:
            year: The year.
            month: The month.

        Returns:
            Dictionary with daily_counts, total, etc.
        """
        return self._report_service.get_monthly_appointments(year, month)

    # ── Doctor Reports ─────────────────────────────────────────

    @require_role(Role.ADMIN)
    def get_doctor_workload(self) -> List[Dict[str, Any]]:
        """Get appointment counts per doctor.

        Returns:
            List of workload records.
        """
        return self._report_service.get_doctor_workload()

    @require_role(Role.ADMIN)
    def get_department_statistics(self) -> List[Dict[str, Any]]:
        """Get appointment counts per department.

        Returns:
            List of department statistics.
        """
        return self._report_service.get_department_statistics()

    # ── Patient Reports ────────────────────────────────────────

    @require_role(Role.ADMIN)
    def get_patient_count(self) -> int:
        """Get total patient count.

        Returns:
            Total number of patients.
        """
        return self._report_service.get_patient_count()

    @require_role(Role.ADMIN)
    def get_patient_gender_distribution(self) -> List[Dict[str, Any]]:
        """Get patient counts by gender.

        Returns:
            List of dicts with 'gender' and 'count'.
        """
        return self._report_service.get_patient_gender_distribution()

    @require_role(Role.ADMIN)
    def get_recent_patients(self, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recently registered patients.

        Args:
            limit: Number of records.

        Returns:
            List of recent patients.
        """
        return self._report_service.get_recent_patients(limit)

    # ── Analytics ──────────────────────────────────────────────

    @require_role(Role.ADMIN)
    def get_peak_hours(self) -> List[Dict[str, Any]]:
        """Get appointment distribution by hour.

        Returns:
            List of dicts with 'hour' and 'count'.
        """
        return self._report_service.get_peak_hours()

    @require_role(Role.ADMIN)
    def get_status_counts(self) -> List[Dict[str, Any]]:
        """Get appointment counts by status.

        Returns:
            List of dicts with 'status' and 'count'.
        """
        return self._report_service.get_appointment_status_report()

    # ── Clinical Records ───────────────────────────────────────

    @require_role(Role.DOCTOR)
    def get_patient_visits(self, patient_id: str) -> List[Dict[str, Any]]:
        """Get visit history for a patient.

        Args:
            patient_id: The patient ID.

        Returns:
            List of visit records.
        """
        return self._clinical_service.get_patient_visits(patient_id)

    @require_role(Role.DOCTOR)
    def get_visit(self, visit_id: int) -> Optional[Dict[str, Any]]:
        """Get a visit record with prescriptions and reports.

        Args:
            visit_id: The visit ID.

        Returns:
            Visit record or None.
        """
        return self._clinical_service.get_visit(visit_id)

    @require_role(Role.DOCTOR)
    def create_visit(self, data: Dict[str, Any]) -> Tuple[bool, str, Optional[int]]:
        """Create a new visit record.

        Args:
            data: Visit fields.

        Returns:
            Tuple of (success, message, visit_id_or_None).
        """
        return self._clinical_service.create_visit(data)

    @require_role(Role.DOCTOR)
    def update_visit(
        self, visit_id: int, data: Dict[str, Any]
    ) -> Tuple[bool, str]:
        """Update a visit record.

        Args:
            visit_id: The visit ID.
            data: Fields to update.

        Returns:
            Tuple of (success, message).
        """
        return self._clinical_service.update_visit(visit_id, data)

    @require_role(Role.DOCTOR)
    def add_prescription(
        self, visit_id: int, data: Dict[str, Any]
    ) -> Tuple[bool, str, Optional[int]]:
        """Add a prescription to a visit.

        Args:
            visit_id: The visit ID.
            data: Prescription fields.

        Returns:
            Tuple of (success, message, prescription_id_or_None).
        """
        return self._clinical_service.add_prescription(visit_id, data)

    @require_role(Role.DOCTOR)
    def add_report(
        self, visit_id: int, data: Dict[str, Any]
    ) -> Tuple[bool, str, Optional[int]]:
        """Add a test report to a visit.

        Args:
            visit_id: The visit ID.
            data: Report fields.

        Returns:
            Tuple of (success, message, report_id_or_None).
        """
        return self._clinical_service.add_report(visit_id, data)

    # ── Analytics ──────────────────────────────────────────────

    @require_role(Role.ADMIN)
    def get_patient_registrations(
        self, start_date: date, end_date: date,
    ) -> List[Dict[str, Any]]:
        """Get patient registration counts by date.

        Args:
            start_date: Range start.
            end_date: Range end.

        Returns:
            List of dicts with 'registration_date' and 'count'.
        """
        return self._report_service.get_patient_registrations(start_date, end_date)

    @require_role(Role.ADMIN)
    def get_cancellation_rate(
        self, start_date: date, end_date: date,
    ) -> List[Dict[str, Any]]:
        """Get cancellation rate per date within a range.

        Args:
            start_date: Range start.
            end_date: Range end.

        Returns:
            List of dicts with date, total, cancelled, rate.
        """
        return self._report_service.get_cancellation_rate(start_date, end_date)

    @require_role(Role.ADMIN)
    def get_analytics_data(
        self, start_date: date, end_date: date,
    ) -> Dict[str, Any]:
        """Get all analytics data for a date range in one call.

        Args:
            start_date: Range start.
            end_date: Range end.

        Returns:
            Dictionary with all analytics keys.
        """
        return self._report_service.get_analytics_data(start_date, end_date)

    # ── Export ─────────────────────────────────────────────────

    # ── Analytics Export ───────────────────────────────────────

    @require_role(Role.ADMIN)
    def export_analytics_pdf(
        self, start_date: date, end_date: date, filepath: str,
        chart_paths=None,
    ) -> Tuple[bool, str]:
        """Export analytics dashboard to PDF.

        Calls the service to gather data, then delegates to
        the PDF export helper.  Chart snapshots can be embedded
        by passing PNG file paths.

        Args:
            start_date: Report start date.
            end_date: Report end date.
            filepath: Destination file path.
            chart_paths: Optional list of PNG paths to embed.

        Returns:
            Tuple of (success, message).
        """
        return self._report_service.export_analytics_pdf(
            start_date, end_date, filepath, chart_paths=chart_paths,
        )

    @require_role(Role.ADMIN)
    def export_analytics_excel(
        self, start_date: date, end_date: date, filepath: str,
    ) -> Tuple[bool, str]:
        """Export analytics data to Excel.

        Args:
            start_date: Report start date.
            end_date: Report end date.
            filepath: Destination file path.

        Returns:
            Tuple of (success, message).
        """
        return self._report_service.export_analytics_excel(
            start_date, end_date, filepath,
        )

    @require_role(Role.ADMIN)
    def prepare_export(
        self, report_type: str, user_id: Optional[int] = None, **kwargs: Any
    ) -> Tuple[List[str], List[List[Any]], str]:
        """Prepare data for PDF/Excel export and log the action.

        Args:
            report_type: Type of report to generate.
            user_id: The user performing the export.
            **kwargs: Additional parameters.

        Returns:
            Tuple of (headers, rows, title).
        """
        headers, rows, title = self._report_service.prepare_export_data(report_type, **kwargs)
        if user_id:
            self._audit_service.log_export(user_id, report_type)
        return headers, rows, title

    # ── Unified Report Generation ─────────────────────────────

    @require_role(Role.ADMIN)
    def generate_report(
        self,
        report_type: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> List[Dict[str, Any]]:
        """Generate a report of the given type, returning row dicts.

        This is the single entry-point used by the ReportsView.  It
        dispatches to the appropriate service method and normalises
        the output into a flat list of dictionaries suitable for
        display in a Treeview.

        Args:
            report_type: One of 'daily_appointments',
                'monthly_appointments', 'doctor_workload',
                'department_stats', 'patient_demographics',
                'cancellation_rate', 'peak_hours'.
            start_date: Optional start date (for date-range reports).
            end_date: Optional end date (for date-range reports).

        Returns:
            List of row dictionaries.
        """
        if report_type == "daily_appointments":
            return self.get_daily_appointments(start_date)
        elif report_type == "monthly_appointments":
            data = self.get_monthly_appointments(
                year=start_date.year if start_date else None,
                month=start_date.month if start_date else None,
            )
            return data.get("daily_counts", [])
        elif report_type == "doctor_workload":
            return self.get_doctor_workload()
        elif report_type == "department_stats":
            return self.get_department_statistics()
        elif report_type == "patient_demographics":
            return self.get_patient_gender_distribution()
        elif report_type == "cancellation_rate":
            if start_date and end_date:
                return self.get_cancellation_rate(start_date, end_date)
            return []
        elif report_type == "peak_hours":
            return self.get_peak_hours()
        else:
            return []

    @require_role(Role.ADMIN)
    def export_report_pdf(
        self,
        filepath: str,
        headers: List[str],
        rows: List[List[Any]],
        title: str,
        report_type: str,
    ) -> Tuple[bool, str]:
        """Export arbitrary report data to a PDF file.

        Args:
            filepath: Destination file path.
            headers: Column header labels.
            rows: Data rows (list of lists).
            title: Report title.
            report_type: The report type key (for logging).

        Returns:
            Tuple of (success, message).
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            )
            from reportlab.lib.styles import getSampleStyleSheet

            doc = SimpleDocTemplate(
                filepath, pagesize=landscape(A4),
                rightMargin=36, leftMargin=36,
                topMargin=36, bottomMargin=36,
            )
            styles = getSampleStyleSheet()
            story: list = []

            story.append(Paragraph(title, styles["Title"]))
            story.append(Spacer(1, 12))

            if rows:
                table_data = [headers] + [
                    [str(cell) for cell in row] for row in rows
                ]
                num_cols = len(headers)
                col_width = (landscape(A4)[0] - 72) / max(num_cols, 1)
                t = Table(table_data, colWidths=[col_width] * num_cols)
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#F8F9FA")]),
                ]))
                story.append(t)
            else:
                story.append(Paragraph(
                    "No data available for this report.", styles["Normal"],
                ))

            doc.build(story)
            logger.info("Report PDF exported: %s", filepath)
            return True, f"PDF exported to {filepath}"
        except Exception as e:
            logger.error("Report PDF export failed: %s", e)
            return False, f"PDF export failed: {e}"

    @require_role(Role.ADMIN)
    def export_report_excel(
        self,
        filepath: str,
        headers: List[str],
        rows: List[List[Any]],
        title: str,
        report_type: str,
    ) -> Tuple[bool, str]:
        """Export arbitrary report data to an Excel file.

        Args:
            filepath: Destination file path.
            headers: Column header labels.
            rows: Data rows (list of lists).
            title: Report title.
            report_type: The report type key (for logging).

        Returns:
            Tuple of (success, message).
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel sheet name limit

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="2C3E50", end_color="2C3E50", fill_type="solid",
            )
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Auto-size columns
            for col_idx, h in enumerate(headers, 1):
                max_len = max(
                    len(str(h)),
                    max((len(str(row[col_idx - 1])) for row in rows), default=0),
                )
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = min(max_len + 2, 40)

            wb.save(filepath)
            logger.info("Report Excel exported: %s", filepath)
            return True, f"Excel exported to {filepath}"
        except Exception as e:
            logger.error("Report Excel export failed: %s", e)
            return False, f"Excel export failed: {e}"
