"""Export service — PDF and Excel report generation.

These helpers are intentionally kept in the service layer so that
both controllers and GUI modules can invoke them without creating
an architecture inversion (service → GUI dependency).
"""
from __future__ import annotations

import logging
import os
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


def export_analytics_pdf(
    data: Dict[str, Any],
    start_date: date,
    end_date: date,
    filepath: str,
    chart_paths: Optional[List[str]] = None,
) -> Tuple[bool, str]:
    """Export analytics data and chart images to a PDF file.

    Uses reportlab to lay out summary tables and embedded chart images.

    Args:
        data: Analytics data dictionary.
        start_date: Report start date.
        end_date: Report end date.
        filepath: Destination file path.
        chart_paths: Optional list of PNG file paths to embed.

    Returns:
        Tuple of (success, message).
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=72)
        styles = getSampleStyleSheet()
        story = []

        # Title
        story.append(Paragraph(
            f"Analytics Report: {start_date.isoformat()} to {end_date.isoformat()}",
            styles["Title"],
        ))
        story.append(Spacer(1, 12))

        # Summary stats
        daily = data.get("daily_appointments", {})
        total_appts = daily.get("total", 0)
        total_patients = sum(r.get("count", 0) for r in data.get("patient_registrations", []))
        story.append(Paragraph(f"Total Appointments: {total_appts}", styles["Heading2"]))
        story.append(Paragraph(f"New Patients: {total_patients}", styles["Heading2"]))
        story.append(Spacer(1, 12))

        # Daily appointments table
        story.append(Paragraph("Daily Appointments", styles["Heading2"]))
        daily_counts = daily.get("daily_counts", [])
        if daily_counts:
            table_data = [["Date", "Appointments"]]
            for d in daily_counts:
                table_data.append([str(d.get("appointment_date", "")), d.get("count", 0)])
            t = Table(table_data, colWidths=[2*inch, 1.5*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t)
            story.append(Spacer(1, 12))

        # Doctor workload table
        story.append(Paragraph("Doctor Workload", styles["Heading2"]))
        workload = data.get("doctor_workload", [])
        if workload:
            table_data = [["Doctor", "Department", "Appointments"]]
            for w in workload[:15]:
                table_data.append([
                    w.get("doctor_name", ""),
                    w.get("department_name", ""),
                    w.get("appointment_count", 0),
                ])
            t = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.2*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (2, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t)
            story.append(Spacer(1, 12))

        # Department breakdown
        story.append(Paragraph("Department Statistics", styles["Heading2"]))
        dept_stats = data.get("department_stats", [])
        if dept_stats:
            table_data = [["Department", "Appointments"]]
            for d in dept_stats:
                table_data.append([d.get("department_name", ""), d.get("count", 0)])
            t = Table(table_data, colWidths=[2.5*inch, 1.5*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t)
            story.append(PageBreak())

        # Embed chart images
        if chart_paths:
            story.append(Paragraph("Chart Snapshots", styles["Heading2"]))
            story.append(Spacer(1, 8))
            for path in chart_paths:
                if os.path.exists(path):
                    img = Image(path, width=6*inch, height=3*inch)
                    story.append(img)
                    story.append(Spacer(1, 12))

        doc.build(story)
        logger.info("PDF exported: %s", filepath)
        return True, f"PDF exported to {filepath}"
    except Exception as e:
        logger.error("PDF export failed: %s", e)
        return False, f"PDF export failed: {e}"


def export_analytics_excel(
    data: Dict[str, Any],
    start_date: date,
    end_date: date,
    filepath: str,
) -> Tuple[bool, str]:
    """Export analytics data to an Excel workbook with multiple sheets.

    Args:
        data: Analytics data dictionary.
        start_date: Report start date.
        end_date: Report end date.
        filepath: Destination file path.

    Returns:
        Tuple of (success, message).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Helper to write a sheet
        def _write_sheet(ws, title: str, headers: List[str], rows: List[List[Any]]) -> None:
            ws.title = title
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="2C3E50", end_color="2C3E50", fill_type="solid",
            )
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            ws.column_dimensions["A"].width = 20
            if len(headers) > 1:
                ws.column_dimensions["B"].width = 18
            if len(headers) > 2:
                ws.column_dimensions["C"].width = 18

        # Sheet 1: Daily Appointments
        ws1 = wb.active
        daily = data.get("daily_appointments", {})
        daily_counts = daily.get("daily_counts", [])
        _write_sheet(ws1, "Daily Appointments",
                     ["Date", "Appointments"],
                     [[str(d.get("appointment_date", "")), d.get("count", 0)]
                      for d in daily_counts])

        # Sheet 2: Patient Registrations
        ws2 = wb.create_sheet()
        registrations = data.get("patient_registrations", [])
        _write_sheet(ws2, "Patient Registrations",
                     ["Date", "Registrations"],
                     [[str(r.get("registration_date", "")), r.get("count", 0)]
                      for r in registrations])

        # Sheet 3: Doctor Workload
        ws3 = wb.create_sheet()
        workload = data.get("doctor_workload", [])
        _write_sheet(ws3, "Doctor Workload",
                     ["Doctor", "Department", "Appointments"],
                     [[w.get("doctor_name", ""), w.get("department_name", ""),
                       w.get("appointment_count", 0)] for w in workload])

        # Sheet 4: Department Statistics
        ws4 = wb.create_sheet()
        dept_stats = data.get("department_stats", [])
        _write_sheet(ws4, "Departments",
                     ["Department", "Appointments"],
                     [[d.get("department_name", ""), d.get("count", 0)]
                      for d in dept_stats])

        # Sheet 5: Cancellation Rate
        ws5 = wb.create_sheet()
        cancel_data = data.get("cancellation_rate", [])
        _write_sheet(ws5, "Cancellations",
                     ["Date", "Total", "Cancelled", "Rate %"],
                     [[str(c.get("appointment_date", "")),
                       c.get("total", 0),
                       c.get("cancelled", 0),
                       c.get("rate", 0)] for c in cancel_data])

        # Sheet 6: Peak Hours
        ws6 = wb.create_sheet()
        peak = data.get("peak_hours", [])
        _write_sheet(ws6, "Peak Hours",
                     ["Hour", "Appointments"],
                     [[f"{int(p.get('hour', 0)):02d}:00", p.get("count", 0)]
                      for p in peak])

        wb.save(filepath)
        logger.info("Excel exported: %s", filepath)
        return True, f"Excel exported to {filepath}"
    except Exception as e:
        logger.error("Excel export failed: %s", e)
        return False, f"Excel export failed: {e}"
